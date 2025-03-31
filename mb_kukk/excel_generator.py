import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from utilities import config, logger, timer_decorator
from datetime import datetime
import sqlite3
import win32com.client
from pathlib import Path


class DailyReport:

    def __init__(self, date, do_print=False):
        self.exceptions = []
        self.date = date
        # Формируем имя файла на основе даты
        self.file_name = (Path() /
                          f'reports\\{self.date.strftime("%d.%m.%Y")}_report.xlsx').absolute()
        # формирование отчета при условии его существования
        if not self.report_exists():
            self.fetch_data()
            self.create_template(neft_in=len(self.neft_in),
                                 neft_out=len(self.neft_out))

        self.print_out_report(print_out=do_print)

    # создание отчета
    @timer_decorator
    def create_template(self, neft_in=1, neft_out=1):
        # Вычисляем количество строк для сырья и продуктов
        try:
            month_dict = {1: 'января', 2: 'февраля', 3: 'марта', 4: 'апреля', 5: 'мая', 6: 'июня',
                          7: 'июля', 8: 'августа', 9: 'сентября', 10: 'октября', 11: 'ноября', 12: 'декабря'}

            in_height = neft_in + 6  # Количество строк для сырья
            out_height = neft_in + neft_out + 8  # Количество строк для продуктов

            # Создаем новую книгу и активный лист
            wb = openpyxl.Workbook()
            ws = wb.active

            # Шаблоны оформления
            align_center = Alignment(
                horizontal='center', vertical='center', wrap_text=True)
            bordered = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
            side_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'))

            # Устанавливаем ширину столбцов
            column_widths = {'A': 30, 'B': 8, 'C': 10, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10,
                             'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 10}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # Применяем выравнивание и границы ко всем ячейкам
            for row in ws[f'B1:O{out_height+1+len(self.urov)}']:
                for cell in row:
                    cell.alignment = align_center

            for row in ws[f'A5:O{out_height}']:
                for cell in row:
                    cell.border = side_border

            # Основной заголовок Excel файла
            ws['A1'] = config['excel_title']
            ws.merge_cells('A1:O2')
            for row in ws['A1:O2']:
                for cell in row:
                    cell.font = Font(bold=True, size=16)
                    cell.border = bordered

            # Наименование сырья и продуктов
            ws['A3'] = 'Наим-ние сырья и продуктов'
            ws['A3'].font = Font(bold=True)

            ws.merge_cells('A3:A4')
            for row in ws['A3:A4']:
                for cell in row:
                    cell.border = bordered

            # Плотность
            ws['B3'] = 'Плотн. Кг/м3'
            ws.merge_cells('B3:B4')
            for row in ws['B3:B4']:
                for cell in row:
                    cell.border = bordered

            # Вахты
            cells = ['0-8', '8-16', '16-0', 'Сутки', 'План', 'Месяц']
            ranges = ['C3:D3', 'E3:F3', 'G3:H3', 'I3:K3', 'L3', 'M3:O3']
            for c, r in zip(cells, ranges):
                cell = r.split(':')[0] if ':' in r else r
                ws[cell] = c
                if ':' in r:
                    ws.merge_cells(r)
                for row in ws['C3:O3']:
                    for cell in row:
                        cell.border = bordered

            # Единицы измерения и проценты
            units = ['м3', 'тн', 'м3', 'тн', 'м3', 'тн',
                     'м3', 'тн', '%', '(%/т)', 'м3', 'тн', '%']
            columns = ['C4', 'D4', 'E4', 'F4', 'G4', 'H4',
                       'I4', 'J4', 'K4', 'L4', 'M4', 'N4', 'O4']
            for unit, col in zip(units, columns):
                ws[col] = unit
                ws[col].border = bordered

            # Заголовок для сырья
            ws['A5'] = 'Взято'
            ws['A5'].font = Font(bold=True)

            # Итоговая строка для сырья
            ws[f'A{in_height}'] = 'Всего'
            ws[f'A{in_height}'].font = Font(bold=True)
            for row in ws[f'A{in_height}:O{in_height}']:
                for cell in row:

                    cell.border = bordered

            # Заголовок для продуктов
            ws[f'A{in_height + 1}'] = 'Получено'

            # Итоговая строка для продуктов
            ws[f'A{out_height}'] = 'Итого'
            ws[f'A{out_height}'].font = Font(bold=True)
            for row in ws[f'A{out_height}:O{out_height}']:
                for cell in row:
                    cell.border = bordered

            # Строка потерь
            ws[f'A{out_height+1}'] = 'Потери'
            ws[f'A{out_height+1}'].font = Font(bold=True)
            for row in ws[f'A{out_height+1}:O{out_height+1}']:
                for cell in row:
                    cell.border = bordered

            # заполнение имен колонок сырья

            # вычисление суммы сырья за сутки и за месяц для вычисления процентов
            sum_of_day_in = sum([i['sum_day_m'] for i in self.neft_in])
            sum_of_month_in = sum([i['all_mass'] for i in self.neft_in])

            for name, letter in zip(('ima', 'P_nom',
                                    f'Sz_N{self.date.day}_1V',
                                    f'M_N{self.date.day}_1V',
                                    f'Sz_N{self.date.day}_2V',
                                    f'M_N{self.date.day}_2V',
                                    f'Sz_N{self.date.day}_3V',
                                    f'M_N{self.date.day}_3V',
                                    'sum_day_vol', 'sum_day_m',
                                     'sum_day_m',
                                     'all_vol', 'all_mass',
                                     'all_mass'),
                                    ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N', 'O')):

                # sum_col хранит сумму колонки
                sum_col = 0
                for data, row in zip(self.neft_in, ws[f'{letter}6:{letter}{in_height-1}']):

                    for cell in row:
                        if letter == 'K' and sum_of_day_in != 0:
                            cell.value = (
                                data['sum_day_m'] / sum_of_day_in) * 100
                        elif letter == 'O' and sum_of_month_in != 0:
                            cell.value = (
                                data['all_mass'] / sum_of_month_in) * 100
                        else:
                            cell.value = data[name]

                        # суммирование sum_col если колонка в нужном диапазоне
                        # исключаются колонки плотности и наименования
                        if cell.column in range(3, 16):
                            sum_col += cell.value
                # вставка суммы (sum_col) колонки в ряд "Всего"
                if letter in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N', 'O'):
                    ws[f'{letter}{in_height}'].value = sum_col
            # ___________________________________________________________________________

            # вычисление суммы продукта за сутки и за месяц для вычисления процентов
            sum_of_day_in = sum([i['sum_day_m'] for i in self.neft_out])
            sum_of_month_in = sum([i['all_mass'] for i in self.neft_out])

            for name, letter in zip(('ima', 'P_nom',
                                    f'Sz_N{self.date.day}_1V',
                                    f'M_N{self.date.day}_1V',
                                    f'Sz_N{self.date.day}_2V',
                                    f'M_N{self.date.day}_2V',
                                    f'Sz_N{self.date.day}_3V',
                                    f'M_N{self.date.day}_3V',
                                    'sum_day_vol', 'sum_day_m',
                                     'sum_day_m',
                                     'all_vol', 'all_mass',
                                     'all_mass'),

                                    ('A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N', 'O')):

                # sum_col хранит сумму колонки
                sum_col = 0
                for data, row in zip(self.neft_out, ws[f'{letter}{in_height+2}:{letter}{out_height-1}']):

                    for cell in row:
                        if letter == 'K' and sum_of_day_in != 0:
                            cell.value = (
                                data['sum_day_m'] / sum_of_day_in) * 100
                        elif letter == 'O' and sum_of_month_in != 0:
                            cell.value = (
                                data['all_mass'] / sum_of_month_in) * 100
                        else:
                            cell.value = data[name]
                        # суммирование sum_col если колонка в нужном диапазоне
                        # исключаются колонки плотности и наименования
                        if cell.column in range(3, 16):
                            sum_col += cell.value

                # вставка суммы (sum_col) колонки в ряд "Итого"
                if letter in ('C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'M', 'N', 'O'):
                    ws[f'{letter}{out_height}'].value = sum_col
                    # вычисление потерь и вставка в строку потери
                    ws[f'{letter}{out_height+1}'].value = ws[f'{letter}{out_height}'].value - \
                        ws[f'{letter}{out_height}'].value
            # ___________________________________________________________________________
            # отображение уровней в отчете при их наличии
            if self.urov:
                for name, letter in zip(('ima',
                                         f'Sz_N{self.date.day}_1V',
                                         f'Sz_N{self.date.day}_2V',
                                         f'Sz_N{self.date.day}_3V',),
                                        ('A', 'C', 'E', 'G')):

                    for data, row in zip(self.urov, ws[f'{letter}{out_height+2}:{letter}{out_height+2+len(self.urov)}']):
                        for cell in row:
                            cell.value = data[name]
                            cell.font = Font(bold=True)

            # вставка даты отчета
            date = self.date
            ws[f'A{out_height+4+len(self.urov)}'].value = f'{date.day} {month_dict[date.month]} {date.year}'

            # применение горизонтальной ориентации
            ws.page_setup.orientation = 'landscape'
            # Вписать все столбцы на одну страницу
            ws.page_setup.fitToPage = True

            # Сохранение файла
            wb.save(self.file_name)
            logger.info(f"Отчет {self.file_name.name} создан")

        except Exception as er:
            tb = er.__traceback__
            logger.error(f'error_on_line: {tb.tb_lineno} | {er} ')

    # формирование срезов данных о сырье и продукте и создание атрибутов
    # и разделение на neft_in и neft_out
    def fetch_data(self) -> None:

        cur_day = self.date
        try:
            with sqlite3.connect(config['local_db']['db_path']) as con:

                # проверка требуется ли печать из архива
                if self.date.month < datetime.now().month or self.date.year < datetime.now().year:
                    table_name = f'Archive_{self.date.year}_{self.date.month:02}'
                else:
                    table_name = 'table1'

                # весь объем за месяц
                all_vol = '+'.join([f'Sz_N{d}_{w}V' for d in range(1, cur_day.day + 1)
                                    for w in range(1, 4)])

                # вся масса за месяц
                all_mass = '+'.join([f'M_N{d}_{w}V' for d in range(1, cur_day.day + 1)
                                    for w in range(1, 4)])

                # cумма вахт обёма и массы за день
                sum_day_vol = '+'.join(
                    [f'Sz_N{cur_day.day}_{w}V' for w in range(1, 4)])
                sum_day_m = '+'.join(
                    [f'M_N{cur_day.day}_{w}V' for w in range(1, 4)])

                # три вахты массы и объёма за день
                cur_day_watches_vol = ','.join((sum_day_vol.split('+')))
                cur_day_watches_mass = ','.join((sum_day_m.split('+')))

                con.row_factory = sqlite3.Row
                curs = con.cursor()

                in_sql = f'''select Ima, Shifr, prizn, ed, P_nom,
                {cur_day_watches_vol},{sum_day_vol} as sum_day_vol,
                {all_vol} as all_vol,
                {cur_day_watches_mass},{sum_day_m} as sum_day_m,
                {all_mass} as all_mass from {table_name}'''

                curs.execute(in_sql)
                rows = curs.fetchall()

                # Разделение на сырье и продукт
                self.neft_in = [
                    row for row in rows if row['prizn'].startswith('NEFT_IN')]
                self.neft_out = [
                    row for row in rows if row['prizn'].startswith('NEFT_OUT')]
                self.urov = [
                    row for row in rows if row['prizn'].startswith('UROV')]

        except Exception as er:
            tb = er.__traceback__
            logger.error(f'error_on_line: {tb.tb_lineno} | {er} ')
            raise

        finally:
            con.close()

    # распечатка отчета при условии, что атрибут print_out = True
    def print_out_report(self, print_out):
        try:
            if print_out:
                copies = config['excel_copies']
                # Создание экземпляра Excel
                excel = win32com.client.Dispatch("Excel.Application")
                # Открытие файла
                workbook = excel.Workbooks.Open(self.file_name)
                worksheet = workbook.Worksheets(1)
                for _ in range(copies):
                    worksheet.PrintOut()
                workbook.Close(SaveChanges=False)
                excel.Quit()
                logger.info(f"Печать {self.file_name.name} завершена")
        except Exception as er:
            tb = er.__traceback__
            logger.error(f'error_on_line: {tb.tb_lineno} | {er} ')

    # метод для проверки существования отчета
    def report_exists(self):
        return self.file_name.exists()
